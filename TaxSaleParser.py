from xlwt import Workbook
import openpyxl
import docx

class TaxSaleParser:

    #option: file name string (must be excel) containing the excel with county option info
    def __init__(self, option: str) -> None:
        self.text = self.readtxt("../Source/source.docx")
        self.option = option
        self.output_file_name = "result"
        self.args_dict = {}
        self.info_list = []
        self.parsed_dict = {}

    def find_args(self):
        wb = openpyxl.load_workbook(self.option)
        sheet = wb.active
        fields = self.get_fields(sheet)
        for idx, field in enumerate(fields):
            self.args_dict[field] = self.get_start_and_end(sheet, idx)
            if (idx == 0):
                self.keyword = self.args_dict[field][0]
        print(self.args_dict)

    def get_fields(self, sheet):
        # i = 1
        # field_list = []
        # print(sheet[f"A{i}"])
        # value = sheet[f'A{i}']
        # while (not(value == '')):
        #     field_list.append(value)
        #     i += 1
        #     value = sheet[f'A{i}']
        # return field_list

        field_list = []
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if (row[0] is not None):
                field_list.append(row[0])
        print(field_list)
        return field_list

    def get_start_and_end(self, sheet, i):
        i += 2  #To account for first row of headers
        return (sheet[f"B{i}"].value, sheet[f"C{i}"].value)

    def parse_info(self):
        for field in self.args_dict:
            values = []
            start_text = self.args_dict[field][0]
            end_text = self.args_dict[field][1]
            for info in self.info_list:
                value = self.find_text(info, start_text, end_text, 0)
                if (value is None):
                    value = ''
                values.append(value)
            self.parsed_dict[field] = values 
        print(self.parsed_dict)

    def revise_info(self):
        for idx, info in enumerate(self.info_list):
            self.info_list[idx] = self.keyword + info
    
    def generate_excel(self):
        workbook = Workbook()
        sheet1 = workbook.add_sheet("Tax Sales")
        self.parsed_dict["Info"] = self.info_list
        for idx, field in enumerate(self.args_dict):
            sheet1.write(0, idx, field)
        # sheet1.write(0, 0, 'File #')
        # sheet1.write(0, 1, 'Map/Parcel Number')
        # sheet1.write(0, 2, 'Owner')
        # sheet1.write(0, 3, 'Address or Location')
        # sheet1.write(0, 4, 'Additional info')
        row_offset = 1
        col = 0
        for field in self.parsed_dict:
            for row, val in enumerate(self.parsed_dict[field]):
                sheet1.write(row + row_offset, col, val)
            col += 1
        workbook.save(f"../Result/{self.output_file_name}.xls")

    def find_text(self, text, start_text, end_text, offset):
        index_of_address = text.find(start_text)
        if index_of_address == -1:
            return None
        start_index = index_of_address + len(start_text) + offset
        if (end_text == "" or end_text is None):
            end_index = len(text) - 1
        else:
            end_index = start_index + text[start_index::].find(end_text)
        return text[start_index:end_index].strip()

    def split_info(self):
        # print(self.keyword)
        self.info_list = self.text.split(self.keyword)
        self.info_list.pop(0)
        self.revise_info()
        # print(self.info_list[0])

    def readtxt(self, filename):
        doc = docx.Document(filename)
        fullText = []
        for para in doc.paragraphs:
            fullText.append(para.text)
        return ' '.join(fullText)

parser = TaxSaleParser("../source/options.xlsx",)
# parser.retrieve_parcel()
parser.find_args()
parser.split_info()
parser.parse_info()
parser.generate_excel()
# parser.retrieve_file()
# parser.retrieve_parcel()
# parser.retrieve_owner()
# parser.retrieve_address()
# parser.generate_excel()